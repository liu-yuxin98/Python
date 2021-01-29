import os
os.system("pip install openpyxl")

import openpyxl
wb=openpyxl.Workbook()
ws=wb.active

ws['A'+str(1)].value="miles"
ws['B'+str(2)].value="kilometers"
for i in range(2,10):
    ws['A'+str(i)].value=i
    ws['B'+str(i)].value=ws['A'+str(i)].value*1.67

wb.save('5.4.xlsx')